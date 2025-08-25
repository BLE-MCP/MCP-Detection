# -*- coding: utf-8 -*-
import os
import json
import time
import argparse
import pandas as pd
import re
from dataclasses import dataclass
from collections import defaultdict
from typing import Any, Dict, List, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ========================= Adjustable parameters =========================
MAX_FLOWS_PER_FILE  = 200_000
HEARTBEAT_EVERY     = 100
PREVIEW_MAX_ITEMS   = 5
TRACE_PREVIEW_MAX   = 220
FLOW_BUILD_FILTER   = False     
FAST_NO_TRACE       = True

DETERMINISTIC_PAIR_LIMIT_PER_VAR = 100 

# ========================= Basic configuration/tools =========================
SPLIT_PATTERN = re.compile(r"[._:/\\-]+")
ENCRYPTION_KEYWORDS = [
    "encrypt", "encryption", "cipher", "md5", "crypto", "private_key", "public_key",
    "key_iv", "hashlib.md5", "md5.new", "md5.sum", "md5.create", "digest::md5", "cc_md5",
    "md5", "md5::compute", "md5::new", "md5::digest", "cryptojs.md5",
    "crypto.createhash('md5')", "crypto.createhash(\"md5\")", "sparkmd5",
    "hashlib.sha1", "sha1.new", "sha1.sum", "sha1.computehash",
    "digest::sha1", "sha1", "sha1::digest", "sha1::new", "cc_sha1", "cryptojs.sha1", "cryptojs",
    "crypto.createhash('sha1')", "require('sha1')", "messagedigest.getinstance", "crypto.createhash(\"sha1\")",
    "des", "rsa", "aes", "decrypt", "hashlib.new"
]
STRICT_MATCH_KEYWORDS = {"des", "rsa", "aes"}
IMPORTANT_FIELDS = {"api", "function", "arguments", "args", "params", "method", "call", "operation"}
LOW_CONFIDENCE_FIELDS = {"trace", "location", "line", "file", "_trace", "produced_as"}
FALSE_POSITIVE_API_PREFIXES = ["uuid.newsha1", "uuid.sha1", "uuid.v5", "uuid.v3"]
FALSE_POSITIVE_FIELD_CONTEXTS = ["commit.sha", "commitinfo.sha", "version.sha", "git.sha", "metadata.sha"]

# CBC/IV
IV_KEY_NAMES = {"iv", "init_vector", "initialization_vector", "nonce"}
CBC_IV_POSITION_HINTS = {
    "createcipheriv": 2, "openssl_encrypt": 4, "cipher.newcbcencrypter": 1,
    "evp_encryptinit_ex": 4, "evp_decryptinit_ex": 4, "aes.new": 2,
}


API_KEY_NAME_HINTS = {
    "api_key","apikey","api-key","apikeyid","apikey","x-api-key","x_api_key",
    "authorization","auth","token","access_token","bearer",
    "openai_api_key","anthropic_api_key","gemini_api_key",
    "google_api_key","groq_api_key","cohere_api_key",
    "huggingface_token","hf_token","hf_access_token","x-authorization","x-authorization-token","apiKey"
}
HEADER_KEY_HINTS = {"authorization","x-api-key","x-authorization","x-authorization-token","api-key"}

API_KEY_VALUE_REGEXES = [
    re.compile(r"sk-[A-Za-z0-9]{32,64}$"),
    re.compile(r"sk-ant-[A-Za-z0-9]{24,70}$"),
    re.compile(r"sk-(live|test)-[A-Za-z0-9]{24,64}$"),
    re.compile(r"gsk_[A-Za-z0-9]{24,80}$"),
    re.compile(r"hf_[A-Za-z0-9]{30,80}$"),
    re.compile(r"AIza[0-9A-Za-z_\-]{30,60}$"),
]
_PLACEHOLDER_HINTS = {
    "<token>","{token}","{api_key}","<api_key>","your api key",
    "bearer token for authentication","example","sample","placeholder"
}
_NOISE_CHARS = set("/\\:?=&.@,;{}()[]<>|\"'` +")

def _redact(token: str) -> str:
    if not isinstance(token, str):
        token = str(token)
    t = token.strip()
    if len(t) <= 12:
        return t
    return f"{t[:6]}…{t[-4:]}"

def _is_probably_secret_name(name: str) -> bool:
    n = (name or "").lower()
    if n in API_KEY_NAME_HINTS:
        return True
    return any(k in n for k in ["api_key","apikey","token","authorization","x-api-key"])

def _normalize_token_value(val: str) -> str:
    v = (val or "").strip()
    if (v.startswith('"') and v.endswith('"')) or (v.startswith("'") and v.endswith("'")):
        v = v[1:-1].strip()
    low = v.lower()
    if low.startswith("bearer "):
        v = v.split(" ", 1)[1].strip()
    return v

def _looks_like_llm_api_key_value(val: str) -> bool:
    if not isinstance(val, str):
        return False
    v = _normalize_token_value(val)
    low = v.lower()
    if any(h in low for h in _PLACEHOLDER_HINTS):
        return False
    if any(ch in _NOISE_CHARS for ch in v):
        return False
    if len(v) < 24:
        return False
    for rgx in API_KEY_VALUE_REGEXES:
        if rgx.fullmatch(v):
            return True
    return False

def _check_hardcoded_api_key_in_args(args: dict) -> Tuple[bool, str]:
    def _val_is_key(name, v) -> Tuple[bool, str]:
        if isinstance(v, dict) and v.get("type") == "constant":
            raw = v.get("value")
            if isinstance(raw, str) and _is_probably_secret_name(name) and _looks_like_llm_api_key_value(raw):
                show = _normalize_token_value(raw)
                return True, f"{name}='{_redact(show)}'"
        return False, ""
    for k_, v_ in (args or {}).items():
        name_l = str(k_).lower()
        hit, ev = _val_is_key(name_l, v_)
        if hit:
            return True, ev
        if isinstance(v_, dict) and v_.get("type") == "dict_literal":
            for hk, hv in (v_.get("value") or {}).items():
                hk_l = str(hk).lower()
                if hk_l in HEADER_KEY_HINTS or _is_probably_secret_name(hk_l):
                    if isinstance(hv, dict) and hv.get("type") == "constant":
                        raw = hv.get("value")
                        if isinstance(raw, str) and _looks_like_llm_api_key_value(raw):
                            show = _normalize_token_value(raw)
                            return True, f"{k_}.{hk}='{_redact(show)}'"
    return False, ""

TAINT_SOURCE_APIS = {
    "input", "raw_input", "sys.argv", "argparse.parse_args", "os.getenv",
    "flask.request.args.get", "flask.request.form.get", "request.getparameter",
    "http.get", "requests.get", "requests.post", "urllib.request.urlopen",
    "express.req.query", "express.req.body", "koa.ctx.query", "koa.ctx.request.body",
    "java.servlet.request.getparameter", "scanf", "gets", "readline",
}
RNG_APIS = {"random.random", "random.randint", "secrets.token_bytes", "os.urandom", "np.random.rand"}
FIXED_SEED_APIS = {"random.seed", "np.random.seed", "torch.manual_seed"}

class TaintInfo:
    __slots__ = ("tainted", "labels", "is_constant", "const_value")
    def __init__(self, tainted=False, labels=None, is_constant=False, const_value=None):
        self.tainted = tainted
        self.labels = set(labels or [])
        self.is_constant = is_constant
        self.const_value = const_value
    def merge(self, other: "TaintInfo"):
        return TaintInfo(
            tainted=self.tainted or other.tainted,
            labels=self.labels | other.labels,
            is_constant=self.is_constant and other.is_constant,
            const_value=self.const_value if self.is_constant else (other.const_value if other.is_constant else None),
        )

class TaintEngine:
    def __init__(self, ir_data):
        self.ir = ir_data or {}
        self.var_map = build_variable_map(ir_data)
        self.cache = {}
        self.fixed_seed_values = []
        self._pre_scan_fixed_seed()
    def _pre_scan_fixed_seed(self):
        for file_entry in (self.ir.get("files") or []):
            for call in (file_entry.get("calls") or []):
                api = (call.get("api") or "").lower()
                if any(seed in api for seed in FIXED_SEED_APIS):
                    args = call.get("arguments", {}) or {}
                    items = list(args.items())
                    if items:
                        arg = items[0][1]
                        if is_constant(arg):
                            self.fixed_seed_values.append(get_constant_value(arg))
    def get_info_from_arg(self, arg_ir: dict) -> TaintInfo:
        if not isinstance(arg_ir, dict):
            return TaintInfo()
        t = arg_ir.get("type")
        v = arg_ir.get("value")
        if t == "constant":
            return TaintInfo(tainted=False, is_constant=True, const_value=v)
        if t == "unknown":
            traced = trace_to_constant(v, self.var_map)
            if traced and traced.get("type") == "constant":
                return TaintInfo(tainted=False, is_constant=True, const_value=traced.get("value"))
            return TaintInfo(tainted=True, labels={"unknown"})
        if t == "function_return":
            src_api = str(v).lower()
            if any(src_api.endswith(s) or src_api == s for s in TAINT_SOURCE_APIS):
                return TaintInfo(tainted=True, labels={"ext_input"})
            if src_api in RNG_APIS:
                labels = {"rng"}
                if self.fixed_seed_values:
                    labels.add("rng_fixed_seed")
                return TaintInfo(tainted=False, labels=labels)
        if t == "list_literal":
            info = TaintInfo(is_constant=True, const_value=[])
            for item in (v or []):
                info = info.merge(self.get_info_from_arg(item))
            return info
        if t == "dict_literal":
            info = TaintInfo(is_constant=True, const_value={})
            for vv in (v or {}).values():
                info = info.merge(self.get_info_from_arg(vv))
            return info
        if isinstance(v, str) and v in self.var_map:
            return self.get_info_from_var(v)
        return TaintInfo()
    def get_info_from_var(self, var_name: str, visited=None) -> TaintInfo:
        if var_name in self.cache:
            return self.cache[var_name]
        if visited is None:
            visited = set()
        if var_name in visited:
            return TaintInfo()
        visited.add(var_name)
        arg = self.var_map.get(var_name)
        if not arg:
            return TaintInfo(tainted=True, labels={"unknown_var"})
        info = self.get_info_from_arg(arg)
        self.cache[var_name] = info
        return info

def parse_arguments():
    p = argparse.ArgumentParser("Detect crypto misuses with flows-first tracing (path-aware)")
    p.add_argument("input_dir")
    p.add_argument("--excel", required=True)
    p.add_argument("--output", required=True)
    return p.parse_args()

def is_false_positive_api(api_str):
    api = (api_str or "").lower()
    if any(api.startswith(p) for p in FALSE_POSITIVE_API_PREFIXES):
        return True
    if re.search(r'\bdes\.(strip|lstrip|rstrip|split|replace|lower|upper|title|join)\b', api):
        return True
    return False

def is_false_positive_field(text):
    text = (text or "").lower()
    return any(text.startswith(p) for p in FALSE_POSITIVE_FIELD_CONTEXTS)

def is_constant(arg_ir):
    return isinstance(arg_ir, dict) and arg_ir.get("type") == "constant"

def get_constant_value(arg_ir):
    return arg_ir.get("value") if is_constant(arg_ir) else None

def build_variable_map(ir_data):
    var_map = {}
    for fi, file_entry in enumerate(ir_data.get("files") or []):
        for ci, call in enumerate(file_entry.get("calls") or []):
            for arg_name, arg in (call.get("arguments") or {}).items():
                if isinstance(arg, dict) and "produced_as" in arg:
                    arg["_ctx"] = {"fi": fi, "ci": ci, "arg": arg_name}
                    var_map[arg["produced_as"]] = arg
            for key in ("return", "result", "out"):
                r = call.get(key)
                if isinstance(r, dict) and "produced_as" in r:
                    r["_ctx"] = {"fi": fi, "ci": ci, "arg": key}
                    var_map[r["produced_as"]] = r
            if isinstance(call.get("returns"), list):
                for idx, r in enumerate(call["returns"]):
                    if isinstance(r, dict) and "produced_as" in r:
                        r["_ctx"] = {"fi": fi, "ci": ci, "arg": f"returns[{idx}]"}
                        var_map[r["produced_as"]] = r
    return var_map

def trace_to_constant(var_name, var_map, visited=None):
    if visited is None:
        visited = set()
    if var_name in visited:
        return None
    visited.add(var_name)
    arg = var_map.get(var_name)
    if arg is None:
        return None
    if is_constant(arg):
        return arg
    val = arg.get("value", "")
    if isinstance(val, str) and val in var_map:
        return trace_to_constant(val, var_map, visited)
    return None


DES_ALGO_TOKEN_RE = re.compile(r'(?i)\bDES(?:EDE3|EDE)?(?:[-/](?:CBC|ECB|CFB|OFB))?(?:/PKCS5Padding)?\b')
CRYPTO_API_HINTS = (
    "cipher.getinstance", "openssl_encrypt", "evp_", "cipher.",
    "cryptojs", "descbc::new_from_slices", "des::", "newcipher", "newtripledescipher"
)
DES_API_BOUNDARY_REGEXES = [
    re.compile(r'(?i)\bdes\.new\b'),
    re.compile(r'(?i)\bdes\.create\b'),
    re.compile(r'(?i)\bdes\.newcipher\b'),
    re.compile(r'(?i)\bdes\.newtripledescipher\b'),
    re.compile(r'(?i)\bcryptojs\.des\b'),
]
def _api_looks_crypto(api: str) -> bool:
    a = (api or "").lower()
    return any(h in a for h in CRYPTO_API_HINTS)

def _extract_des_evidence_with_arg(api: str, args):
    api_l = (api or "").lower()
    if not _api_looks_crypto(api_l):
        return None, None, None, None
    for k_, a in (args or {}).items():
        if isinstance(a, dict) and a.get("type") == "constant":
            v = a.get("value")
            if isinstance(v, str):
                m = DES_ALGO_TOKEN_RE.search(v)
                if m:
                    tok = m.group(0)
                    return tok.lower(), tok, a, k_
        if isinstance(a, dict) and a.get("type") == "dict_literal":
            for kk, vv in (a.get("value") or {}).items():
                if isinstance(vv, dict) and vv.get("type") == "constant":
                    v = vv.get("value")
                    if isinstance(v, str):
                        m = DES_ALGO_TOKEN_RE.search(v)
                        if m:
                            tok = m.group(0)
                            return tok.lower(), tok, vv, kk
    return None, None, None, None

def _extract_des_evidence(api: str, args):
    key_tok, match_tok, _, _ = _extract_des_evidence_with_arg(api, args)
    return key_tok, match_tok

def is_weak_sha_reference(token, field, full_text):
    return (
        token == "sha" and (
            is_false_positive_api(full_text) or
            is_false_positive_field(full_text) or
            "uuid" in full_text or
            "commit" in full_text or
            "version" in full_text
        )
    )

try:
    import networkx as nx
except Exception:
    nx = None

def _get_call_produced_var(call: dict):
    """If there is no real product, a synthetic product node is generated to ensure that every call can fall on the graph."""
    for key in ("return", "result", "out"):
        r = call.get(key)
        if isinstance(r, dict) and r.get("produced_as"):
            return r.get("produced_as")
    if isinstance(call.get("returns"), list):
        for r in call["returns"]:
            if isinstance(r, dict) and r.get("produced_as"):
                return r.get("produced_as")
    if call.get("produced_as"):
        return call.get("produced_as")
    # —— Compositing Node：CALL@file#function:line
    loc = call.get("location", {}) or {}
    tag = f"CALL@{loc.get('file','?')}#{loc.get('function','?')}:{loc.get('line','?')}"
    return tag

def _build_data_flow_graph(all_calls, extra_edges: List[Tuple[str, str]] = None):
    """ Build the variable-level flow graph: the real edge arg_var -> produced_var; and merge the extra_edges (produced_var -> produced_var). """
    if nx is None:
        return None
    G = nx.DiGraph()

    for ci, call in enumerate(all_calls):
        prod = _get_call_produced_var(call)
        args = call.get("arguments", {}) or {}
        if prod:
            G.add_node(prod, call=call, ci=ci)
        for arg in args.values():
            if isinstance(arg, dict):
                if arg.get("type") == "unknown" and isinstance(arg.get("value"), str):
                    val = arg.get("value")
                    if not G.has_node(val):
                        G.add_node(val, call=None)
                    if prod:
                        G.add_edge(val, prod)
                else:
                    val = arg.get("value")
                    if isinstance(val, str):
                        if not G.has_node(val):
                            G.add_node(val, call=None)
                        if prod:
                            G.add_edge(val, prod)

    if extra_edges:
        for a, b in extra_edges:
            if not a or not b:
                continue
            if not G.has_node(a):
                G.add_node(a, call=None)
            if not G.has_node(b):
                G.add_node(b, call=None)
            G.add_edge(a, b)  # latent edge
    return G

def _trace_taint(var, G, visited=None):
    if G is None:
        return []
    if visited is None:
        visited = set()
    if var in visited or var not in G:
        return []
    visited.add(var)
    taint_calls = []
    for succ in G.successors(var):
        call_info = G.nodes[succ].get("call")
        if call_info:
            taint_calls.append(call_info)
        taint_calls.extend(_trace_taint(succ, G, visited))
    return taint_calls

# Data flow reachability: whether the call product can reach the sensitive convergence
def _is_sensitive_sink(api: str) -> bool:
    a = (api or "").replace(" ", "").lower()
    SENSITIVE_SINK_APIS = [
        "print","logger","logging","console.log","http","request","post","send","fetch",
        "os.system","exec","subprocess","write","dump","save","open","response","return",
        "json.dump","ctx.body","res.json"
    ]
    return any(a.endswith(s) or a == s for s in SENSITIVE_SINK_APIS)

def _leaks_to_sensitive_sink(call, graph) -> bool:
    if graph is None:
        return False
    v = _get_call_produced_var(call)
    if not v or v not in graph:
        return False
    try:
        import networkx as nx
        for succ in nx.descendants(graph, v):
            c = graph.nodes[succ].get("call")
            if c and _is_sensitive_sink((c.get("api") or "")):
                return True
    except Exception:
        return False
    return False

HASH_APIS = {
    "md5": [
        "hashlib.md5", "md5.new", "md5.sum", "md5.create", "digest::md5", "cc_md5",
        "md5", "md5::compute", "md5::new", "md5::digest", "cryptojs.md5",
        "crypto.createhash('md5')", "crypto.createhash(\"md5\")", "sparkmd5",
        "messagedigest.getinstance", "hashlib.new"
    ],
    "sha1": [
        "hashlib.sha1", "sha1.new", "sha1.sum", "sha1.computehash", "digest::sha1", "sha1",
        "sha1::sha1", "sha1::digest", "sha1::new", "cc_sha1", "cryptojs.sha1",
        "crypto.createhash('sha1')", "require('sha1')", "messagedigest.getinstance",
        "crypto.createhash(\"sha1\")", "hashlib.new"
    ]
}

SAFE_CONTEXT_KEYWORDS = ["file", "read", "path", "checksum", "verify", "validate", "compare", "digest"]
SENSITIVE_KEYWORDS = ["token", "sign", "signature", "auth", "apikey", "secret"]
SENSITIVE_CONTEXT_NAMES = ["auth", "sign", "token", "header", "signature", "apikey", "login"]
DATA_ARG_HINTS = {"data","text","message","plaintext","input","buf","payload","content","body","value"}

def _is_hash_function(api: str, hash_type: str) -> bool:
    a = (api or "").replace(" ", "").lower()
    return any(a.endswith(h) or a == h for h in HASH_APIS[hash_type])

def _argument_values_flat(arguments):
    vals = []
    for arg in (arguments or {}).values():
        if isinstance(arg, dict):
            t = arg.get("type")
            v = arg.get("value")
            if t in ("constant", "function_return"):
                vals.append(str(v).lower())
            elif t == "list_literal":
                for item in (arg.get("value") or []):
                    if isinstance(item, dict) and item.get("type") == "constant":
                        vals.append(str(item.get("value")).lower())
            elif t == "dict_literal":
                for _, vv in (arg.get("value") or {}).items():
                    if isinstance(vv, dict):
                        vals.append(str(vv.get("value", "")).lower())
    return vals

def _pick_hash_data_arg(args: dict):
    items = list((args or {}).items())
    if not items:
        return None, None
    for k, v in items:
        if str(k).lower() in DATA_ARG_HINTS:
            return k, v
    k0, v0 = items[0]
    return k0, v0

def _is_file_hashing_context(call):
    api = (call.get("api") or "").lower()
    args = call.get("arguments", {}) or {}
    const_vals = _argument_values_flat(args)
    return any(kw in api for kw in SAFE_CONTEXT_KEYWORDS) or \
           any(kw in val for kw in SAFE_CONTEXT_KEYWORDS for val in const_vals)

def _hash_alg_from_args(api: str, args: dict) -> str:
    a = (api or "").lower()
    if ("messagedigest.getinstance" in a) or ("hashlib.new" in a):
        items = list((args or {}).items())
        if items:
            _, v = items[0]
            if isinstance(v, dict) and v.get("type") == "constant":
                val = str(v.get("value","")).strip().lower().replace("-", "")
                if val == "md5": return "md5"
                if val in ("sha1","sha"): return "sha1"
    return ""

def _detect_hash_usage(all_calls, taint_engine=None, taint_strict=False, graph=None, extra_edges=None):
    if graph is None:
        graph = _build_data_flow_graph(all_calls, extra_edges=extra_edges)
    issues = []
    G = graph

    for idx, call in enumerate(all_calls):
        api = (call.get("api") or "").lower()
        args = call.get("arguments", {}) or {}
        loc = call.get("location", {}) or {}
        produced_var = _get_call_produced_var(call)
        func_name = str(loc.get("function", "")).lower()
        file_name = str(loc.get("file", "")).lower()

        data_arg_name, data_arg_ir = _pick_hash_data_arg(args)
        taint_ok = True
        if taint_engine is not None and data_arg_ir is not None and taint_strict:
            tinfo = taint_engine.get_info_from_arg(data_arg_ir)
            taint_ok = (tinfo.tainted or ("ext_input" in tinfo.labels))

        alg_by_arg = _hash_alg_from_args(api, args)

        for htype in ("md5", "sha1"):
            if _is_hash_function(api, htype):
                if alg_by_arg and alg_by_arg != htype:
                    continue
                sensitive_by_args = any(key in val for val in _argument_values_flat(args) for key in SENSITIVE_KEYWORDS)
                sensitive_by_name = (any(w in func_name for w in SENSITIVE_CONTEXT_NAMES) or
                                     any(w in file_name for w in SENSITIVE_CONTEXT_NAMES))
                if taint_ok and (sensitive_by_args or sensitive_by_name):
                    issues.append({
                        "hash_type": htype,
                        "rule_violation": f"{htype.upper()} used in sensitive context",
                        "hash_function": api,
                        "call_index": idx
                    })

                if _is_file_hashing_context(call):
                    continue

                if produced_var:
                    taint_chain = _trace_taint(produced_var, G) if G is not None else []
                    for leak_call in taint_chain:
                        if _is_sensitive_sink(leak_call.get("api", "")):
                            issues.append({
                                "hash_type": htype,
                                "rule_violation": f"{htype.upper()} result leaked via trace".replace("UPPER","upper"),
                                "hash_function": api,
                                "leaked_variable": produced_var,
                                "leak_via": leak_call.get("api"),
                                "call_index": idx
                            })
                            break
    return issues

def find_encryption_in_json(json_data):
    hits = {
        "keywords": set(),
        "fields": set(),
        "matched_strings": set(),
        "has_api_crypto": False,
        "has_argument_key": False,
        "has_strong_encryption": False,
    }
    def scan(text, field, is_key=False):
        if not isinstance(text, str):
            return
        text_l = text.lower()
        field_l = field.lower()
        if field_l == "api" and is_false_positive_api(text_l):
            return
        if is_false_positive_field(text_l):
            return
        tokens = SPLIT_PATTERN.split(text_l)
        if field_l == "api":
            if any(kw in text_l for kw in ["encrypt","decrypt","cipher","sign","crypto","hash","hmac"]):
                hits["has_api_crypto"] = True
        if field_l in {"arguments","args","params"} and any(k in text_l for k in ["key","secret","signature","hmac"]):
            hits["has_argument_key"] = True
        for word in ENCRYPTION_KEYWORDS:
            for token in tokens:
                if word in STRICT_MATCH_KEYWORDS or word in {"encrypt","decrypt","cipher","aes","des","rsa","private_key","public_key","createcipher","createcipheriv"}:
                    if token == word and not is_weak_sha_reference(token, field_l, text_l):
                        hits["has_strong_encryption"] = True
                        hits["keywords"].add(word)
                        hits["fields"].add(field_l)
                        hits["matched_strings"].add(text)
                else:
                    if word in token:
                        if word in {"encrypt","decrypt","cipher","aes","des","rsa","private_key","public_key","createcipher","createcipheriv"}:
                            hits["has_strong_encryption"] = True
                        hits["keywords"].add(word)
                        hits["fields"].add(field_l)
                        hits["matched_strings"].add(text)
    def search(obj, parent_key=None):
        if isinstance(obj, dict):
            for k, v in obj.items():
                key_lower = str(k).lower()
                if key_lower in LOW_CONFIDENCE_FIELDS:
                    continue
                if key_lower in IMPORTANT_FIELDS or parent_key in IMPORTANT_FIELDS:
                    scan(k, key_lower, is_key=True)
                    scan(v, key_lower)
                else:
                    if isinstance(v, (dict, list)):
                        search(v, key_lower)
        elif isinstance(obj, list):
            for item in obj:
                search(item, parent_key)
        elif isinstance(obj, str):
            if parent_key in IMPORTANT_FIELDS:
                scan(obj, parent_key)
    search(json_data)
    signal_count = sum([
        hits["has_api_crypto"],
        hits["has_argument_key"],
        len(hits["keywords"]) >= 1
    ])
    is_encryption = signal_count >= 2
    return is_encryption, list(hits["keywords"]), list(hits["matched_strings"]), hits["has_strong_encryption"]


def _args_text(args) -> str:
    vals = []
    for a in (args or {}).values():
        if isinstance(a, dict):
            v = a.get("value")
            if isinstance(v, str):
                vals.append(v.lower())
            elif isinstance(v, list):
                for item in v:
                    if isinstance(item, dict) and item.get("type") == "constant":
                        vals.append(str(item.get("value")).lower())
            elif isinstance(v, dict):
                for vv in v.values():
                    if isinstance(vv, dict) and vv.get("type") == "constant":
                        vals.append(str(vv.get("value")).lower())
    return " ".join(vals)

def _has_cbc(api: str, args) -> bool:
    hay = f"{api} {_args_text(args)}".lower()
    return "cbc" in hay

def _extract_cbc_iv_info_with_arg(api: str, args, taint):
    api_l = (api or "").lower()
    iv_info = None
    iv_evidence = None
    iv_arg_ir = None
    iv_arg_name = None

    for k_, v_ in (args or {}).items():
        k_l = str(k_).lower()
        if any(alias in k_l for alias in IV_KEY_NAMES):
            if isinstance(v_, dict):
                info = taint.get_info_from_arg(v_)
                iv_info = info
                iv_arg_ir = v_
                iv_arg_name = k_
                iv_evidence = f"{k_}={repr(info.const_value) if info.is_constant else '[non-const]'}"
                break
    if iv_info is None:
        for hint, pos in CBC_IV_POSITION_HINTS.items():
            if hint in api_l:
                items = list((args or {}).items())
                if items:
                    if pos >= len(items):
                        pos = len(items) - 1
                    if 0 <= pos < len(items):
                        kpos, apos = items[pos]
                        info = taint.get_info_from_arg(apos)
                        iv_info = info
                        iv_arg_ir = apos
                        iv_arg_name = kpos
                        iv_evidence = f"pos{pos}={repr(info.const_value) if info.is_constant else '[non-const]'}"
                break
    if iv_info is None:
        arg_text = _args_text(args)
        if "modes.cbc" in arg_text or "blockmode: cbc" in arg_text or "cbc(" in arg_text:
            for k_, v_ in (args or {}).items():
                if isinstance(v_, dict) and v_.get("type") == "dict_literal":
                    for kk, vv in (v_.get("value") or {}).items():
                        if str(kk).lower() in IV_KEY_NAMES and isinstance(vv, dict):
                            info = taint.get_info_from_arg(vv)
                            iv_info = info
                            iv_arg_ir = vv
                            iv_arg_name = kk
                            iv_evidence = f"nested.{kk}={repr(info.const_value) if info.is_constant else '[non-const]'}"
                            break
                if iv_info is not None:
                    break
    if iv_info is None:
        arg_text = _args_text(args)
        if "ivparameterspec" in arg_text or "ivparameterspec" in api_l:
            items = list((args or {}).items())
            if items:
                k0, v0 = items[0]
                info = taint.get_info_from_arg(v0)
                iv_info = info
                iv_arg_ir = v0
                iv_arg_name = k0
                iv_evidence = f"IvParameterSpec({repr(info.const_value) if info.is_constant else '[non-const]'})"
    return iv_info, iv_evidence, iv_arg_ir, iv_arg_name

def _extract_cbc_iv_info(api: str, args, taint):
    iv_info, iv_ev, _, _ = _extract_cbc_iv_info_with_arg(api, args, taint)
    return iv_info, iv_ev

def detect_crypto_rules(all_calls, taint_engine: TaintEngine, json_data, graph=None, extra_edges=None):
    violations = []
    evidence_keys = []
    evidence_matches = []

    rule1_hit = rule2_hit = rule3_hit = rule4_hit = rule5_hit = False
    rule8_hit = rule9_hit = rule10_hit = False
    rule6_hit = rule7_hit = False  # MD5 / SHA1

    has_mac_function = False

    KEY_USAGE_PATTERNS = [
        {"api": "createcipheriv", "key_pos": 1},
        {"api": "algorithms.aes", "key_pos": 0},
        {"api": "aes.new", "key_pos": 0},
        {"api": "aes.newcipher", "key_pos": 0},
        {"api": "evp_decryptinit_ex", "key_pos": 3},
        {"api": "evp_encryptinit_ex", "key_pos": 3},
        {"api": "secretkeyspec", "key_pos": 0},
        {"api": "aes(", "key_name": "key"},
        {"api": "openssl_encrypt", "key_pos": 2},
        {"api": "secretkeyspec", "key_pos": 0},
    ]
    PBKDF2_APIS = {
        "crypto.pbkdf2": {"salt_pos": 1, "iter_pos": 2},
        "crypto.pbkdf2sync": {"salt_pos": 1, "iter_pos": 2},
        "hashlib.pbkdf2_hmac": {"salt_pos": 2, "iter_pos": 3},
        "pbkdf2.key": {"salt_pos": 1, "iter_pos": 2},
        "rfc2898derivebytes.pbkdf2": {"salt_pos": 1, "iter_pos": 2},
        "pbkdf2.new": {"salt_key": "salt", "iter_key": "iterations"},
        "hash_pbkdf2": {"salt_pos": 2, "iter_pos": 3},
        "pbkdf2_hmac": {"salt_pos": 1, "iter_pos": 2},
    }

    def _get_arg_by_pos(args, pos):
        items = list(args.items())
        if 0 <= pos < len(items):
            return items[pos][1]
        return None

    def _get_named_arg(args, name_lower):
        for k_, v_ in (args or {}).items():
            if str(k_).lower() == name_lower:
                return v_
        return None

    hash_issues = _detect_hash_usage(all_calls, taint_engine=taint_engine, taint_strict=False, graph=graph, extra_edges=extra_edges)
    first_md5 = next((i for i in hash_issues if i["hash_type"] == "md5"), None)
    first_sha1 = next((i for i in hash_issues if i["hash_type"] == "sha1"), None)

    for call in all_calls:
        api = (call.get("api") or "").lower()
        args = (call.get("arguments", {}) or {})

        if (not rule6_hit) and first_md5:
            violations.append("Rule 6: " + first_md5["rule_violation"])
            evidence_keys.append("md5")
            evm = first_md5["hash_function"]
            if "leak_via" in first_md5 and first_md5["leak_via"]:
                evm += f" → {first_md5['leak_via']}"
            evidence_matches.append(evm)
            rule6_hit = True

        if (not rule7_hit) and first_sha1:
            violations.append("Rule 7: " + first_sha1["rule_violation"])
            evidence_keys.append("sha1")
            evm = first_sha1["hash_function"]
            if "leak_via" in first_sha1 and first_sha1["leak_via"]:
                evm += f" → {first_sha1['leak_via']}"
            evidence_matches.append(evm)
            rule7_hit = True

        if not rule1_hit and any(k in api for k in ["cipher","aes","des","encrypt","pydes","modes","algorithm"]):
            for k_, v_ in args.items():
                key_str = str(k_).lower()
                val_str = str(v_.get("value", "")).lower() if isinstance(v_, dict) else ""
                if "ecb" in key_str or "ecb" in val_str:
                    violations.append("Rule 1: Detected ECB mode")
                    evidence_keys.append("ecb_mode")
                    evidence_matches.append(key_str if "ecb" in key_str else val_str)
                    if _leaks_to_sensitive_sink(call, graph):
                        violations.append("Rule 1: Detected ECB mode (result leaked via trace)")
                        evidence_keys.append("ecb_mode")
                        evidence_matches.append((call.get("api") or "").lower())
                    rule1_hit = True
                    break

        if not rule2_hit and ("cbc" in (api + " " + _args_text(args)).lower()):
            iv_info, iv_ev = _extract_cbc_iv_info(api, args, taint_engine)
            if iv_info is not None and iv_info.is_constant:
                violations.append("Rule 2: CBC with constant IV")
                evidence_keys.append("cbc_constant_iv")
                evidence_matches.append(str(iv_ev))
                if _leaks_to_sensitive_sink(call, graph):
                    violations.append("Rule 2: CBC with constant IV (result leaked via trace)")
                    evidence_keys.append("cbc_constant_iv")
                    evidence_matches.append(str(iv_ev))
                rule2_hit = True

        if not rule3_hit:
            for sig in KEY_USAGE_PATTERNS:
                if sig["api"] in api:
                    key_ir = None
                    if "key_pos" in sig:
                        key_ir = _get_arg_by_pos(args, sig["key_pos"])
                    elif "key_name" in sig:
                        key_ir = _get_named_arg(args, sig["key_name"])
                    if key_ir is not None:
                        info = taint_engine.get_info_from_arg(key_ir)
                        if info.is_constant:
                            violations.append("Rule 3: Hardcoded encryption key detected")
                            evidence_keys.append("hardcoded_key")
                            evidence_matches.append(repr(info.const_value))
                            if _leaks_to_sensitive_sink(call, graph):
                                violations.append("Rule 3: Hardcoded encryption key detected (ciphertext leaked via trace)")
                                evidence_keys.append("hardcoded_key")
                                evidence_matches.append(repr(info.const_value))
                            rule3_hit = True
                    break

        if not rule3_hit:
            hit, ev = _check_hardcoded_api_key_in_args(args)
            if hit:
                violations.append("Rule 3: Hardcoded API key detected")
                evidence_keys.append("hardcoded_api_key")
                evidence_matches.append(ev)
                if _leaks_to_sensitive_sink(call, graph):
                    violations.append("Rule 3: Hardcoded API key detected (result leaked via trace)")
                    evidence_keys.append("hardcoded_api_key")
                    evidence_matches.append(ev)
                rule3_hit = True

        for sig, spec in PBKDF2_APIS.items():
            if sig in api:
                if not rule4_hit:
                    salt_ir = None
                    if spec.get("salt_key"):
                        salt_ir = _get_named_arg(args, spec["salt_key"])
                    if salt_ir is None and spec.get("salt_pos") is not None:
                        salt_ir = _get_arg_by_pos(args, spec["salt_pos"])
                    if salt_ir is not None:
                        info = taint_engine.get_info_from_arg(salt_ir)
                        if info.is_constant:
                            violations.append("Rule 10: Constant salt in PBKDF2")
                            evidence_keys.append("pbkdf2_constant_salt")
                            evidence_matches.append(repr(info.const_value))
                            if _leaks_to_sensitive_sink(call, graph):
                                violations.append("Rule 10: Constant salt in PBKDF2 (derived key leaked via trace)")
                                evidence_keys.append("pbkdf2_constant_salt")
                                evidence_matches.append(repr(info.const_value))
                            rule4_hit = True
                if not rule5_hit:
                    iter_ir = None
                    if spec.get("iter_key"):
                        iter_ir = _get_named_arg(args, spec["iter_key"])
                    if iter_ir is None and spec.get("iter_pos") is not None:
                        iter_ir = _get_arg_by_pos(args, spec["iter_pos"])
                    if iter_ir is not None and is_constant(iter_ir):
                        try:
                            iterations = int(get_constant_value(iter_ir))
                            if iterations < 1000:
                                violations.append(f"Rule 3: Low iteration count ({iterations}) in PBKDF2")
                                evidence_keys.append("pbkdf2_low_iterations")
                                evidence_matches.append(str(iterations))
                                if _leaks_to_sensitive_sink(call, graph):
                                    violations.append(f"Rule 3: Low iteration count ({iterations}) in PBKDF2 (derived key leaked via trace)")
                                    evidence_keys.append("pbkdf2_low_iterations")
                                    evidence_matches.append(str(iterations))
                                rule5_hit = True
                        except Exception:
                            pass
                break

        if not rule8_hit and any(seed in api for seed in {"random.seed","np.random.seed","torch.manual_seed"}):
            arg_list = list(args.items())
            if arg_list:
                seed_arg = arg_list[0][1]
                if is_constant(seed_arg):
                    violations.append("Rule 6: Fixed seed in RNG")
                    evidence_keys.append("fixed_seed")
                    evidence_matches.append(str(get_constant_value(seed_arg)))
                    if _leaks_to_sensitive_sink(call, graph):
                        violations.append("Rule 6: Fixed seed in RNG (result leaked via trace)")
                        evidence_keys.append("fixed_seed")
                        evidence_matches.append(str(get_constant_value(seed_arg)))
                    rule8_hit = True

    
        MAC_APIS = [
            "hmac.new","mac.getinstance","crypto.createhmac","generatehmac",
            "cryptojs.hmacsha256","hmacsha256","openssl::hmac.hexdigest",
            "hmac.authenticationcode","hash_hmac","pbkdf2_hmac"
        ]
        if any(mac in api for mac in MAC_APIS):
            has_mac_function = True

        if not rule10_hit:
            if ("cipher.getinstance" in api) or ("openssl_encrypt" in api):
                key_tok, match_tok = _extract_des_evidence(api, args)
                if match_tok:
                    violations.append("Rule 10: DES usage detected")
                    evidence_keys.append(key_tok or "des")
                    evidence_matches.append(match_tok)
                    if _leaks_to_sensitive_sink(call, graph):
                        violations.append("Rule 10: DES usage detected (result leaked via trace)")
                        evidence_keys.append(key_tok or "des")
                        evidence_matches.append(match_tok)
                    rule10_hit = True
            else:
                for rgx in DES_API_BOUNDARY_REGEXES:
                    if rgx.search(api):
                        violations.append("Rule 10: DES usage detected")
                        evidence_keys.append("des_usage")
                        evidence_matches.append(rgx.pattern)
                        if _leaks_to_sensitive_sink(call, graph):
                            violations.append("Rule 10: DES usage detected (result leaked via trace)")
                            evidence_keys.append("des_usage")
                            evidence_matches.append(rgx.pattern)
                        rule10_hit = True
                        break
                if not rule10_hit:
                    key_tok, match_tok = _extract_des_evidence(api, args)
                    if match_tok:
                        violations.append("Rule 10: DES usage detected")
                        evidence_keys.append(key_tok or "des")
                        evidence_matches.append(match_tok)
                        if _leaks_to_sensitive_sink(call, graph):
                            violations.append("Rule 10: DES usage detected (result leaked via trace)")
                            evidence_keys.append(key_tok or "des")
                            evidence_matches.append(match_tok)
                        rule10_hit = True

    is_enc, enc_keywords, enc_matches, has_strong_enc = find_encryption_in_json({
        "files": [{"calls": all_calls}]
    })
    if has_strong_enc and not has_mac_function:
        violations.append("Rule 9: Encryption without MAC detected")
        evidence_keys.append(",".join(enc_keywords) if enc_keywords else "encryption_detected")
        evidence_matches.append(";".join(enc_matches) if enc_matches else "")
        leaked_enc = False
        for c in all_calls:
            api_c = (c.get("api") or "").lower()
            if any(k in api_c for k in ["encrypt","cipher","createcipher","createcipheriv","aes.","rsa.","des.","openssl_encrypt"]):
                if _leaks_to_sensitive_sink(c, graph):
                    leaked_enc = True
                    break
        if leaked_enc:
            violations.append("Rule 9: Encryption without MAC (ciphertext leaked via trace)")
            evidence_keys.append("encryption_no_mac")
            evidence_matches.append("ciphertext->sink")

    return violations, evidence_keys, evidence_matches

ARROW = " → "

def json_path_for(fi, ci, arg=None):
    base = f"files[{fi}].calls[{ci}]"
    return f"{base}.arguments[{repr(str(arg))}]" if arg is not None else base

def _make_ctx_label(ctx, varname=None):
    if not ctx:
        return "?"
    base = f"files[{ctx['fi']}].calls[{ctx['ci']}]"
    if ctx.get("arg") is not None:
        base += f".arguments[{repr(str(ctx['arg']))}]"
    if varname:
        base += f" (VAR {varname})"
    return base

def _find_var_context(json_data, var_name):
    for fi, fe in enumerate(json_data.get("files") or []):
        for ci, call in enumerate(fe.get("calls") or []):
            for arg_name, arg_ir in (call.get("arguments") or {}).items():
                if isinstance(arg_ir, dict) and arg_ir.get("produced_as") == var_name:
                    return {"fi": fi, "ci": ci, "arg": arg_name}
    return None

def _format_const(v):
    if isinstance(v, str):
        return f"CONST '{_redact(v)}'"
    return f"CONST {repr(v)}"

def _trace_nodes_source_to_sink(arg_ir, var_map, json_data, start_ctx=None, visited=None):
    if visited is None:
        visited = set()
    if not isinstance(arg_ir, dict):
        return ["UNKNOWN", _make_ctx_label(start_ctx)]
    t = arg_ir.get("type")
    v = arg_ir.get("value")
    if t == "constant":
        return [_format_const(v), _make_ctx_label(start_ctx)]
    if t == "unknown" and isinstance(v, str):
        if v in visited:
            return [f"UNKNOWN {v} (cycle)", _make_ctx_label(start_ctx)]
        visited.add(v)
        producer_ir = var_map.get(v)
        ctx = _find_var_context(json_data, v)
        if producer_ir is not None and ctx is not None:
            src_nodes = _trace_nodes_source_to_sink(producer_ir, var_map, json_data, start_ctx=ctx, visited=visited)
            if src_nodes:
                src_nodes[-1] = _make_ctx_label(ctx, varname=v)
            return src_nodes + [_make_ctx_label(start_ctx)]
        return [f"UNKNOWN {v}", _make_ctx_label(start_ctx)]
    if t == "function_return":
        return [f"RET {v}", _make_ctx_label(start_ctx)]
    if t in ("dict_literal", "list_literal"):
        return [t.upper(), _make_ctx_label(start_ctx)]
    return ["UNKNOWN", _make_ctx_label(start_ctx)]

def render_trace_arrow(arg_ir, var_map, json_data, start_ctx=None):
    nodes = _trace_nodes_source_to_sink(arg_ir, var_map, json_data, start_ctx=start_ctx)
    cleaned = []
    for n in nodes:
        if n and (not cleaned or cleaned[-1] != n):
            cleaned.append(n)
    return ARROW.join(cleaned)

@dataclass
class Flow:
    fi: int
    ci: int
    api: str
    arg: str
    json_path: str
    source_type: str
    source_value: Any
    trace: str

def _iter_argument_atoms(args: Dict[str, Any]):
    for k, v in (args or {}).items():
        if isinstance(v, dict) and v.get("type") in {"constant","unknown","function_return","dict_literal","list_literal"}:
            yield k, v
            if v.get("type") == "dict_literal":
                for kk, vv in (v.get("value") or {}).items():
                    if isinstance(vv, dict):
                        yield f"{k}.{kk}", vv
            if v.get("type") == "list_literal":
                for idx, vv in enumerate(v.get("value") or []):
                    if isinstance(vv, dict):
                        yield f"{k}[{idx}]", vv

def _source_kind(arg_ir):
    t = (arg_ir or {}).get("type")
    if t == "constant": return "constant"
    if t == "unknown": return "unknown"
    if t == "function_return": return "function_return"
    if t == "dict_literal": return "dict_literal"
    if t == "list_literal": return "list_literal"
    return "unknown_var"

def _looks_relevant_api(api: str) -> bool:
    if not FLOW_BUILD_FILTER:
        return True
    a = (api or "").lower()
    if any(p in a for p in ["log", "logger", "printf", "println", "print", "warn", "error", "geterror", "fmt."]):
        return False
    positive = [
        "encrypt","decrypt","cipher","algorithms.aes","aes.","rsa.","des.","evp_","openssl_encrypt",
        "pbkdf2","hash_pbkdf2","pbkdf2_hmac","hmac","random.seed","np.random.seed","torch.manual_seed",
        "secretkeyspec","cryptojs","newcipher","newcbcencrypter","newcbcdecrypter","newtripledescipher",
        "hashlib.md5","hashlib.sha1","messagedigest.getinstance","hashlib.new","sha1","md5"
    ]
    return any(p in a for p in positive)

def extract_all_flows(json_data, taint_engine) -> List[Flow]:
    flows: List[Flow] = []
    files = (json_data.get("files") or [])
    var_map = taint_engine.var_map
    for fi, fe in enumerate(files):
        calls = (fe.get("calls") or [])
        t_start = time.time()
        for ci, call in enumerate(calls):
            if ci and ci % HEARTBEAT_EVERY == 0:
                elapsed = time.time() - t_start
                rate = ci / max(elapsed, 1e-6)
                eta  = (len(calls)-ci) / max(rate, 1e-6)
                print(f"       … building flows: file[{fi}] {ci}/{len(calls)} calls | {rate:.1f}/s, ETA {eta:.1f}s", flush=True)
            api = (call.get("api") or "")
            if not _looks_relevant_api(api):
                continue
            args = call.get("arguments", {}) or {}
            for arg_name, arg_ir in _iter_argument_atoms(args):
                if FAST_NO_TRACE:
                    trace = ""
                else:
                    trace = render_trace_arrow(arg_ir, var_map, json_data,
                                               start_ctx={"fi": fi, "ci": ci, "arg": arg_name})
                src_kind = _source_kind(arg_ir)
                val = arg_ir.get("value") if isinstance(arg_ir, dict) else None
                if isinstance(val, str) and len(val) > 12:
                    val = _redact(val)
                flows.append(Flow(
                    fi=fi, ci=ci, api=(api or "").lower(), arg=str(arg_name),
                    json_path=f"files[{fi}].calls[{ci}].arguments[{repr(str(arg_name))}]",
                    source_type=src_kind, source_value=val, trace=trace
                ))
                if len(flows) >= MAX_FLOWS_PER_FILE:
                    print(f"       ⚠️ flow cap reached ({MAX_FLOWS_PER_FILE}), truncating this file", flush=True)
                    return flows
    return flows

def index_flows(flows: List[Flow]):
    by_call = defaultdict(list)
    by_key  = {}
    for f in flows:
        by_call[(f.fi, f.ci)].append(f)
        by_key[(f.fi, f.ci, f.arg.lower())] = f
    return by_call, by_key

def _get_arg_key_by_pos(args, pos: int):
    items = list((args or {}).items())
    if not items:
        return None
    if pos >= len(items):
        pos = len(items) - 1
    if 0 <= pos < len(items):
        return str(items[pos][0])
    return None

DATA_ARG_HINTS_WIDE = set(DATA_ARG_HINTS) | {
    "body","json","blob","file","files","stream","bytes","binary",
    "document","record","row","content","payload","message","text","value","input"
}
RESOURCE_ARG_HINTS = {
    "path", "filename", "file", "bucket", "key", "object", "url", "endpoint",
    "topic", "collection", "table", "name", "id", "dest", "destination",
    "dst", "to", "out", "outfile", "out_path"
}
CATEGORY_HINTS = {
    "protect": ["encrypt","cipher","createcipher","createcipheriv","aes","rsa","des","hmac","createhmac","mac","sign","mask","redact","tokenize","anonymize","hash","sha1","md5","sha256","sha-256","sha512"],
    "upload": ["upload","putobject","put_object","s3.put","blob.upload","drive.files.create","files.create","http.post","axios.post","fetch","request.post","request.put","post","put","send","response","res.json","ctx.body","emit","publish","producer.send"],
    "persist": ["store","log","index","write","save","dump","insert","open"],
    "verify": ["verify","verify_mac","verify_signature","verifysignature","hmac.verify","mac.verify","signature.verify"],
    "use": ["decrypt","parse","use","consume","open","read"],
    "derive": ["keygen","generatekey","pbkdf2","unwrap","wrap","derive","kdf","nonce","iv","salt","random","seed"],
    "salt": ["salt"],
    "hash": ["hash","sha1","md5","sha256","sha-256","sha512"],
    "sign": ["sign","createsign","signature"],
    "key_mgmt": ["export","print","save key","savekey","dump key","keystore","serialize","marshal","write key","store key"],
    "mask": ["mask","redact","tokenize","anonymize"],
}
RISK_PAIRS = [
    ("protect", "upload",   "P: Upload-before-Protection possible"),
    ("verify",  "use",      "P: Use/Decrypt-before-Verify possible"),
    ("derive",  "use",      "P: Use-before-Generate/Derive possible"),
    ("salt",    "hash",     "P: Hash-before-Salt possible"),
    ("sign",    "encrypt",  "P: Sign/Encrypt order ambiguous"),
    ("key_mgmt","upload",   "P: Key material release possible"),
    ("mask",    "persist",  "P: Persist-before-Mask/Redact possible"),
]

def _collect_var_refs_from_ir(ir: dict):
    refs = []
    if not isinstance(ir, dict):
        return refs
    t = ir.get("type")
    v = ir.get("value")
    if t == "unknown" and isinstance(v, str):
        refs.append(v.lower())
    elif t == "list_literal":
        for item in (v or []):
            refs.extend(_collect_var_refs_from_ir(item))
    elif t == "dict_literal":
        for _, vv in (ir.get("value") or {}).items():
            refs.extend(_collect_var_refs_from_ir(vv))
    return refs

def extract_var_refs_from_args(args: dict, keys_hint: set = None):
    refs = []
    for k, v in (args or {}).items():
        if keys_hint is not None and str(k).lower() not in keys_hint:
            continue
        refs.extend(_collect_var_refs_from_ir(v))
    return sorted(set(refs))

def extract_resource_constants(args: dict):
    res = []
    for k, v in (args or {}).items():
        k_l = str(k).lower()
        if k_l in RESOURCE_ARG_HINTS and isinstance(v, dict) and v.get("type") == "constant":
            cv = v.get("value")
            if isinstance(cv, (str, int, float)):
                res.append(str(cv).strip().lower())
            continue
        if isinstance(v, dict) and v.get("type") == "dict_literal":
            for kk, vv in (v.get("value") or {}).items():
                if str(kk).lower() in RESOURCE_ARG_HINTS and isinstance(vv, dict) and vv.get("type") == "constant":
                    cv = vv.get("value")
                    if isinstance(cv, (str, int, float)):
                        res.append(str(cv).strip().lower())
    return sorted(set(res))

def _api_in_category(api: str, cat: str) -> bool:
    hints = CATEGORY_HINTS.get(cat, [])
    a = (api or "").lower()
    return any(h in a for h in hints)

def detect_prompt_order_potentials(json_data) -> List[Dict[str, str]]:
    """Potential dependency paths：{'paths': 'files[i].calls[j] || files[x].calls[y]'}"""
    results: List[Dict[str, str]] = []
    files = (json_data.get("files") or [])

    idx_vars: Dict[str, Dict[str, list]] = {cat: defaultdict(list) for cat in CATEGORY_HINTS.keys()}
    idx_res:  Dict[str, Dict[str, list]] = {cat: defaultdict(list) for cat in CATEGORY_HINTS.keys()}

    def _mk_node(fi, ci, api):
        return {"fi": fi, "ci": ci, "api": api, "json_path": f"files[{fi}].calls[{ci}]"}

    for fi, fe in enumerate(files):
        for ci, call in enumerate((fe.get("calls") or [])):
            api = (call.get("api") or "").lower()
            args = call.get("arguments", {}) or {}
            hit_cats = [cat for cat in CATEGORY_HINTS if _api_in_category(api, cat)]
            if not hit_cats:
                continue
            node = _mk_node(fi, ci, api)
            var_refs = extract_var_refs_from_args(args, DATA_ARG_HINTS_WIDE)
            res_tags = extract_resource_constants(args)
            for cat in hit_cats:
                for v in var_refs:
                    idx_vars[cat][v].append(node)
                for r in res_tags:
                    idx_res[cat][r].append(node)

    seen = set()
    def _add_path(a_path, b_path):
        pair = f"{a_path} || {b_path}"
        if pair not in seen:
            seen.add(pair)
            results.append({"paths": pair})

    for a, b, _ in RISK_PAIRS:
        if a not in idx_vars or b not in idx_vars:
            continue
        for v in sorted(set(idx_vars[a].keys()) & set(idx_vars[b].keys())):
            for na in idx_vars[a][v]:
                for nb in idx_vars[b][v]:
                    _add_path(na["json_path"], nb["json_path"])
        for r in sorted(set(idx_res[a].keys()) & set(idx_res[b].keys())):
            for na in idx_res[a][r]:
                for nb in idx_res[b][r]:
                    _add_path(na["json_path"], nb["json_path"])
    return results


def detect_deterministic_dependency_pairs(json_data) -> List[Dict[str, str]]:
    files = (json_data.get("files") or [])
    index = defaultdict(list)  # var -> [{fi,ci,json_path}]
    pairs: List[Dict[str, str]] = []
    for fi, fe in enumerate(files):
        for ci, call in enumerate((fe.get("calls") or [])):
            args = call.get("arguments", {}) or {}
            refs = extract_var_refs_from_args(args, DATA_ARG_HINTS_WIDE)
            if not refs:
                continue
            node = {"fi": fi, "ci": ci, "json_path": f"files[{fi}].calls[{ci}]"}
            for v in refs:
                index[v].append(node)
    for v, nodes in index.items():
        if len(nodes) < 2:
            continue
        seen_local = set()
        cnt = 0
        for i in range(len(nodes)):
            for j in range(i+1, len(nodes)):
                na, nb = nodes[i], nodes[j]
                key = (na["json_path"], nb["json_path"])
                if key in seen_local:
                    continue
                seen_local.add(key)
                pairs.append({"paths": f"{na['json_path']} || {nb['json_path']}"})
                cnt += 1
                if cnt >= DETERMINISTIC_PAIR_LIMIT_PER_VAR:
                    break
            if cnt >= DETERMINISTIC_PAIR_LIMIT_PER_VAR:
                break
    return pairs


_JSON_PATH_FI_CI_RE = re.compile(r"files\[(\d+)\]\.calls\[(\d+)\]")
def _parse_json_path_to_fi_ci(path: str) -> Tuple[int, int]:
    m = _JSON_PATH_FI_CI_RE.search(str(path))
    if not m:
        return -1, -1
    return int(m.group(1)), int(m.group(2))

def _produced_var_index(json_data) -> Dict[Tuple[int,int], str]:
    mp = {}
    for fi, fe in enumerate(json_data.get("files") or []):
        calls = fe.get("calls") or []
        for ci, call in enumerate(calls):
            mp[(fi,ci)] = _get_call_produced_var(call)
    return mp

def build_latent_edges(json_data, relationship_pairs: List[Dict[str,str]]):
    prod = _produced_var_index(json_data)
    global_edges: List[Tuple[str,str]] = []
    per_file_edges: Dict[int, List[Tuple[str,str]]] = defaultdict(list)

    for rel in relationship_pairs or []:
        paths = str(rel.get("paths","")).split("||")
        if len(paths) != 2:
            continue
        a = paths[0].strip()
        b = paths[1].strip()
        fi1, ci1 = _parse_json_path_to_fi_ci(a)
        fi2, ci2 = _parse_json_path_to_fi_ci(b)
        if fi1 < 0 or fi2 < 0:
            continue
        v1 = prod.get((fi1,ci1))
        v2 = prod.get((fi2,ci2))
        if not v1 or not v2 or v1 == v2:
            continue
        if (fi1,ci1) <= (fi2,ci2):
            src, dst = v1, v2
            ffi = fi1 if fi1 == fi2 else -1
        else:
            src, dst = v2, v1
            ffi = fi2 if fi1 == fi2 else -1
        global_edges.append((src, dst))
        if ffi >= 0:
            per_file_edges[ffi].append((src, dst))
    return global_edges, per_file_edges

# ========================= CLI =========================
def _shorten(s: str, limit: int = TRACE_PREVIEW_MAX) -> str:
    if s is None:
        return ""
    s = str(s).replace("\n", " ").strip()
    return (s[:limit] + "…") if len(s) > limit else s

def print_detail_preview(details, max_items: int = PREVIEW_MAX_ITEMS, trace_limit: int = TRACE_PREVIEW_MAX):
    print(f"         ▶ Findings preview (showing {min(len(details), max_items)}/{len(details)}):")
    for d in details[:max_items]:
        rule = d.get("rule", "")
        api = d.get("api", "")
        path = d.get("json_path", "")
        evk  = d.get("evidence_key", "")
        evm  = d.get("evidence_match", "")
        trace = d.get("trace", "")
        trace_disp = _shorten(trace, trace_limit) if trace else "[no trace]"
        print(f"           - {rule}")
        if api:
            print(f"             api : {api}")
        print(f"             path: {path}")
        if evk or evm:
            print(f"             ev  : {evk} | {evm}")
        print(f"             ↳ {trace_disp}")

import gzip, re, json

try:
    import json5
except Exception:
    json5 = None

try:
    import hjson
except Exception:
    hjson = None

_C_LINE_COMMENT = re.compile(r"//.*?$", re.MULTILINE)
_C_BLOCK_COMMENT = re.compile(r"/\*.*?\*/", re.DOTALL)
_TRAILING_COMMA = re.compile(r",\s*([}\]])")
_BAD_NUMS = re.compile(r"\b(NaN|Infinity|-Infinity)\b")
_UNQUOTED_KEY = re.compile(r'([{\[,]\s*)([A-Za-z_][\w\-]*)\s*:', re.MULTILINE)

def _sanitize_step1(text: str) -> str:
    t = _C_BLOCK_COMMENT.sub("", _C_LINE_COMMENT.sub("", text))
    t = _TRAILING_COMMA.sub(r"\1", t)
    t = _BAD_NUMS.sub("null", t)
    t = _UNQUOTED_KEY.sub(lambda m: f'{m.group(1)}"{m.group(2)}":', t)
    return t

def _escape_newlines_in_strings(s: str) -> str:
    out = []
    in_str = False
    quote = None
    esc = False
    for ch in s:
        if in_str:
            if esc:
                out.append(ch); esc = False; continue
            if ch == '\\':
                out.append(ch); esc = True; continue
            if ch == quote:
                in_str = False; out.append(ch); continue
            if ch == '\n' or ch == '\r':
                out.append('\\n'); continue
            if ord(ch) < 0x20:
                out.append('\\u%04x' % ord(ch)); continue
            out.append(ch)
        else:
            out.append(ch)
            if ch == '"' or ch == "'":
                in_str = True; quote = ch; esc = False
    return ''.join(out)

def process_file(filepath):
    try:
        with open(filepath, "rb") as fb:
            raw = fb.read()
    except Exception as e:
        print(f"⚠️  Open failed: {filepath} -> {type(e).__name__}: {e}")
        return None
    if not raw:
        print(f"⚠️  Empty file: {filepath}"); return None
    if raw[:2] == b"\x1f\x8b":
        try:
            raw = gzip.decompress(raw)
        except Exception as e:
            print(f"⚠️  Gzip decompress failed: {filepath} -> {e}")
            return None

    encs = ["utf-8","utf-8-sig","utf-16","utf-16-le","utf-16-be","utf-32","gb18030","latin1"]
    text = None; last_err = None
    for enc in encs:
        try:
            text = raw.decode(enc); break
        except Exception as e:
            last_err = e
    if text is None:
        print(f"⚠️  Decode failed ({type(last_err).__name__}): {filepath} -> {last_err}")
        return None
    t = text.strip()
    if not t:
        print(f"⚠️  Only whitespace: {filepath}")
        return None

    try:
        return json.loads(t)
    except Exception:
        pass

    if json5:
        try:
            return json5.loads(t)
        except Exception:
            pass
    if hjson:
        try:
            return hjson.loads(t)
        except Exception:
            pass

    t2 = _sanitize_step1(t)
    t2 = _escape_newlines_in_strings(t2)
    try:
        return json.loads(t2)
    except Exception:
        if json5:
            try:
                return json5.loads(t2)
            except Exception:
                pass
        if hjson:
            try:
                return hjson.loads(t2)
            except Exception:
                pass
        head = t[:200].replace("\n"," ")
        print(f"⚠️  Invalid JSON: {filepath}")
        print(f"    └─ head preview: {head}")
        return None


def _rules_from_violation_texts(violations: List[str]) -> List[str]:
    rules = set()
    for v in violations or []:
        m = re.search(r"Rule\s*(\d+)", str(v))
        if m:
            rules.add(f"Rule {m.group(1)}")
    return sorted(rules)

def main():
    pd.set_option('display.max_colwidth', None)
    args = parse_arguments()
    print(f"📥 Loading Excel: {args.excel}", flush=True)
    df = pd.read_excel(args.excel)
    df.columns = df.columns.str.strip()

    results_map = {}                 # (project, market) -> (violations, ev_keys, ev_matches)
    potential_paths_map = defaultdict(list)  # (project, market) -> [path_pairs]
    potential_related_rules_map = defaultdict(set)

    json_paths = []
    for root, _, files in os.walk(args.input_dir):
        for file in files:
            if file.endswith(".json"):
                json_paths.append(os.path.join(root, file))

    total = len(json_paths)
    print(f"🔍 Found {total} JSON files under: {args.input_dir}", flush=True)

    for idx, fpath in enumerate(json_paths, start=1):
        print(f"[{idx}/{total}] 📂 Processing: {fpath}", flush=True)
        t0 = time.time()
        json_data = process_file(fpath)
        if json_data is None:
            print(f"⚠️  Skipped invalid or unreadable JSON: {fpath}", flush=True)
            continue
        taint_engine = TaintEngine(json_data)
        print(f"     ⏱ load {time.time()-t0:.2f}s", flush=True)

        
        t1 = time.time()
        flows = extract_all_flows(json_data, taint_engine)
        index_flows(flows)  
        print(f"     ⏱ flows: {len(flows)} in {time.time()-t1:.2f}s", flush=True)

        # === all_calls ===
        all_calls = []
        for file_entry in (json_data.get("files") or []):
            for call in (file_entry.get("calls") or []):
                all_calls.append(call)

        t4 = time.time()
        potentials = detect_prompt_order_potentials(json_data)
        deterministic_pairs = detect_deterministic_dependency_pairs(json_data)
        combined_pairs = []
        combined_pairs.extend(potentials)
        combined_pairs.extend(deterministic_pairs)
        latent_edges_global, _ = build_latent_edges(json_data, combined_pairs)
        print(f"     ⏱ potential pairs {len(combined_pairs)}, latent edges {len(latent_edges_global)} in {time.time()-t4:.2f}s", flush=True)

        # 10 Rules
        t2 = time.time()
        enhanced_graph = _build_data_flow_graph(all_calls, extra_edges=latent_edges_global)
        violations, ev_keys, ev_matches = detect_crypto_rules(all_calls, taint_engine, json_data, graph=enhanced_graph)
        print(f"     ⏱ rules (path-aware on enhanced graph) in {time.time()-t2:.2f}s", flush=True)
   
        project = os.path.basename(fpath)
        market  = os.path.basename(os.path.dirname(fpath))
        key = (project.strip(), market.strip())

        results_map[key] = (violations, ev_keys, ev_matches)

        if combined_pairs:
            for p in combined_pairs:
                potential_paths_map[key].append(p["paths"])
            print(f"     ⚑ Potential dependency paths in ({project}, {market}): {len(combined_pairs)}", flush=True)
            for p in combined_pairs[:min(3, len(combined_pairs))]:
                print(f"         paths: {p['paths']}")

        for r in _rules_from_violation_texts(violations):
            potential_related_rules_map[key].add(r)

        if violations:
            print(f"     ⚠️ Misuse detected in ({project}, {market}): {violations}", flush=True)
            print(f"         ▶ Evidence keywords: {ev_keys}", flush=True)
            print(f"         ▶ Evidence matches : {ev_matches}", flush=True)

    # ========================= Write Excel =========================
    misuse_flags, rule_triggers, evidence_keywords, evidence_matches, misuse_paths = [], [], [], [], []
    potential_flags, potential_paths_col = [], []
    potential_related_rules_col = []

    for _, row in df.iterrows():
        key = (str(row["Project Name"]).strip(), str(row["Market"]).strip())
        vios, keys, mats = results_map.get(key, ([], [], []))
        misuse_flags.append("Yes" if vios else "No")
        rule_triggers.append("; ".join(sorted(set(vios))) if vios else "")
        evidence_keywords.append(", ".join(keys) if keys else "")
        evidence_matches.append("; ".join(mats) if mats else "")
        misuse_paths.append("") 

        ppaths = sorted(set(potential_paths_map.get(key, [])))
        potential_flags.append("")  
        potential_paths_col.append("; ".join(ppaths) if ppaths else "")

        prules = sorted(potential_related_rules_map.get(key, []))
        potential_related_rules_col.append("; ".join(prules) if prules else "")

    df["Misuse Detected"] = misuse_flags
    df["Rule Triggered"] = rule_triggers
    df["Encryption Evidence"] = evidence_keywords
    df["Matched String Evidence"] = evidence_matches
    df["Misuse Paths"] = misuse_paths
    df["Potential Risk Triggered"] = potential_flags
    df["Potential Risk Paths"] = potential_paths_col
    df["Potential Related Rules"] = potential_related_rules_col

    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Summary")

    wb = load_workbook(args.output)
    for ws in wb.worksheets:
        for i, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            header = column_cells[0].value
            if header:
                ws.column_dimensions[get_column_letter(i)].width = min(120, len(str(header)) + 2)
    wb.save(args.output)

    print(f"\n✅ Excel updated (Summary only) and saved to: {args.output}", flush=True)
    print("🎯 Scan complete.", flush=True)

if __name__ == "__main__":
    main()
